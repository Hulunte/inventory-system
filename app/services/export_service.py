from decimal import Decimal
from io import BytesIO

from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models.harvest_entry import HarvestEntry
from app.models.worker import Worker
from app.services.report_service import date_range_to_utc

_DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")


def _safe_text(value):
    if not isinstance(value, str):
        return value
    if value and value[0] in _DANGEROUS_PREFIXES:
        return "'" + value
    return value


def _to_local_naive(dt_utc, tz):
    if dt_utc is None:
        return None
    return dt_utc.astimezone(tz).replace(tzinfo=None)


def _write_header_row(ws, headers):
    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 8, 10), 50)


def generate_harvest_export(start_date, end_date, query_filter=None, tz=None):
    if tz is None:
        tz = current_app.config["HARVEST_TIMEZONE"]

    start_utc, end_utc = date_range_to_utc(start_date, end_date, tz)

    q = (
        db.session.query(
            HarvestEntry.id,
            HarvestEntry.created_at,
            HarvestEntry.weight_kg,
            HarvestEntry.voided,
            HarvestEntry.voided_at,
            HarvestEntry.void_reason,
            HarvestEntry.worker_name_snapshot,
            HarvestEntry.worker_barcode_snapshot,
            HarvestEntry.worker_slot_number_snapshot,
            HarvestEntry.worker_assignment_id,
            HarvestEntry.product_name_snapshot,
            HarvestEntry.rate_per_kg_snapshot,
            HarvestEntry.amount_mxn,
        )
        .filter(
            HarvestEntry.created_at >= start_utc,
            HarvestEntry.created_at < end_utc,
        )
        .order_by(HarvestEntry.created_at.asc(), HarvestEntry.id.asc())
    )

    if query_filter:
        pattern = f"%{query_filter}%"
        q = q.filter(
            db.or_(
                HarvestEntry.worker_name_snapshot.ilike(pattern),
                HarvestEntry.worker_barcode_snapshot.ilike(pattern),
            )
        )

    entries = q.all()

    wb = Workbook()

    ws_mov = wb.active
    ws_mov.title = "Movimientos"

    mov_headers = [
        "ID", "Fecha", "Hora", "Trabajador", "Código", "Cupo",
        "Peso (kg)", "Producto", "Precio/kg", "Importe",
        "Estado", "Fecha y hora de anulación", "Motivo de anulación",
    ]
    _write_header_row(ws_mov, mov_headers)

    for row_idx, entry in enumerate(entries, 2):
        local_created = _to_local_naive(entry.created_at, tz)
        local_voided = _to_local_naive(entry.voided_at, tz)

        ws_mov.cell(row=row_idx, column=1, value=entry.id)

        date_cell = ws_mov.cell(row=row_idx, column=2, value=local_created.date() if local_created else None)
        date_cell.number_format = "YYYY-MM-DD"

        time_cell = ws_mov.cell(row=row_idx, column=3, value=local_created.time() if local_created else None)
        time_cell.number_format = "HH:MM:SS"

        ws_mov.cell(row=row_idx, column=4, value=_safe_text(entry.worker_name_snapshot))
        ws_mov.cell(row=row_idx, column=5, value=_safe_text(entry.worker_barcode_snapshot))

        slot_num = entry.worker_slot_number_snapshot
        ws_mov.cell(row=row_idx, column=6, value=f"Trabajador {slot_num:03d}" if slot_num else None)

        weight_cell = ws_mov.cell(row=row_idx, column=7, value=Decimal(str(entry.weight_kg)))
        weight_cell.number_format = "0.000"

        ws_mov.cell(row=row_idx, column=8, value=entry.product_name_snapshot)

        rate_cell = ws_mov.cell(row=row_idx, column=9, value=entry.rate_per_kg_snapshot)
        if rate_cell.value is not None:
            rate_cell.number_format = "0.00"

        amount_cell = ws_mov.cell(row=row_idx, column=10, value=entry.amount_mxn)
        if amount_cell.value is not None:
            amount_cell.number_format = "0.00"

        ws_mov.cell(row=row_idx, column=11, value="Anulado" if entry.voided else "Vigente")

        if local_voided:
            voided_cell = ws_mov.cell(row=row_idx, column=12, value=local_voided)
            voided_cell.number_format = "YYYY-MM-DD HH:MM:SS"
        else:
            ws_mov.cell(row=row_idx, column=12, value=None)

        ws_mov.cell(row=row_idx, column=13, value=_safe_text(entry.void_reason) if entry.void_reason else None)

    ws_mov.freeze_panes = "A2"
    ws_mov.auto_filter.ref = f"A1:{get_column_letter(len(mov_headers))}{len(entries) + 1}"
    _auto_width(ws_mov)

    ws_res = wb.create_sheet(title="Resumen")

    res_headers = [
        "Trabajador", "Código", "Cupo", "Movimientos vigentes",
        "Peso vigente (kg)", "Movimientos anulados", "Peso anulado (kg)",
    ]
    _write_header_row(ws_res, res_headers)

    worker_summary = {}
    for entry in entries:
        key = (entry.worker_assignment_id, entry.worker_name_snapshot, entry.worker_barcode_snapshot, entry.worker_slot_number_snapshot)
        if key not in worker_summary:
            worker_summary[key] = {
                "vigentes_count": 0,
                "vigentes_weight": Decimal("0.000"),
                "anulados_count": 0,
                "anulados_weight": Decimal("0.000"),
            }
        ws_data = worker_summary[key]
        w = Decimal(str(entry.weight_kg))
        if entry.voided:
            ws_data["anulados_count"] += 1
            ws_data["anulados_weight"] += w
        else:
            ws_data["vigentes_count"] += 1
            ws_data["vigentes_weight"] += w

    total_vig_count = 0
    total_vig_weight = Decimal("0.000")
    total_anul_count = 0
    total_anul_weight = Decimal("0.000")

    worker_data_rows = 0
    row_idx = 2
    for (assignment_id, name, barcode, slot_num), data in sorted(worker_summary.items()):
        ws_res.cell(row=row_idx, column=1, value=_safe_text(name))
        ws_res.cell(row=row_idx, column=2, value=_safe_text(barcode))
        ws_res.cell(row=row_idx, column=3, value=f"Trabajador {slot_num:03d}" if slot_num else None)
        ws_res.cell(row=row_idx, column=4, value=data["vigentes_count"])

        vig_cell = ws_res.cell(row=row_idx, column=5, value=data["vigentes_weight"])
        vig_cell.number_format = "0.000"

        ws_res.cell(row=row_idx, column=6, value=data["anulados_count"])

        anul_cell = ws_res.cell(row=row_idx, column=7, value=data["anulados_weight"])
        anul_cell.number_format = "0.000"

        total_vig_count += data["vigentes_count"]
        total_vig_weight += data["vigentes_weight"]
        total_anul_count += data["anulados_count"]
        total_anul_weight += data["anulados_weight"]

        worker_data_rows += 1
        row_idx += 1

    total_font = Font(bold=True)
    ws_res.cell(row=row_idx, column=1, value="TOTALES").font = total_font
    ws_res.cell(row=row_idx, column=2).font = total_font
    ws_res.cell(row=row_idx, column=3).font = total_font
    ws_res.cell(row=row_idx, column=4, value=total_vig_count).font = total_font

    tv_cell = ws_res.cell(row=row_idx, column=5, value=total_vig_weight)
    tv_cell.number_format = "0.000"
    tv_cell.font = total_font

    ws_res.cell(row=row_idx, column=6, value=total_anul_count).font = total_font

    ta_cell = ws_res.cell(row=row_idx, column=7, value=total_anul_weight)
    ta_cell.number_format = "0.000"
    ta_cell.font = total_font

    res_last_row = 1 + worker_data_rows if worker_data_rows else 1
    ws_res.freeze_panes = "A2"
    ws_res.auto_filter.ref = f"A1:G{res_last_row}"
    _auto_width(ws_res)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_credentials_export():
    from app.models.worker import Worker as WorkerModel
    workers = WorkerModel.query.order_by(WorkerModel.slot_number.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Credenciales"

    headers = ["Número", "Etiqueta", "Código", "Nombre asignado"]
    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, worker in enumerate(workers, 2):
        ws.cell(row=row_idx, column=1, value=worker.slot_number)
        ws.cell(row=row_idx, column=2, value=worker.slot_label)

        barcode_cell = ws.cell(row=row_idx, column=3, value=_safe_text(worker.barcode))
        barcode_cell.number_format = numbers.FORMAT_TEXT

        name_cell = ws.cell(row=row_idx, column=4, value=_safe_text(worker.name) if worker.name else "")
        name_cell.number_format = numbers.FORMAT_TEXT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(workers) + 1}"
    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
