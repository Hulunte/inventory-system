"""Regression tests for worker-slot-assignments feature."""
import io
from datetime import datetime, time, timedelta, timezone
from decimal import Decimal
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook

from app.extensions import db
from app.models.harvest_entry import HarvestEntry
from app.models.product import Product
from app.models.worker import Worker
from app.models.worker_assignment import WorkerAssignment
from app.services.export_service import generate_harvest_export
from app.services.harvest_service import (
    WorkerUnassignedError,
    get_daily_total,
    register_harvest,
)
from app.services.history_service import get_daily_summary, get_worker_entries
from tests.conftest import make_worker, make_worker_with_assignment

HARVEST_TZ = ZoneInfo("America/Mexico_City")


def _ensure_product(db_session):
    product = Product.query.filter_by(name="Naranja").first()
    if product is None:
        product = Product(name="Naranja", rate_per_kg=Decimal("12.50"), active=True)
        db_session.add(product)
        db_session.flush()
    return product


def _make_entry(db_session, worker, assignment, weight_kg, created_at, product=None):
    if product is None:
        product = _ensure_product(db_session)
    entry = HarvestEntry(
        worker_id=worker.id,
        weight_kg=weight_kg,
        product_id=product.id,
        product_name_snapshot=product.name,
        rate_per_kg_snapshot=Decimal(str(product.rate_per_kg)),
        amount_mxn=weight_kg * Decimal(str(product.rate_per_kg)),
        created_at=created_at,
        worker_assignment_id=assignment.id,
        worker_slot_number_snapshot=worker.slot_number,
        worker_barcode_snapshot=worker.barcode,
        worker_name_snapshot=assignment.person_name,
    )
    db_session.add(entry)
    db_session.flush()
    return entry


class TestDailyTotalIsolatedByAssignment:
    def test_total_isolated_per_assignment(self, db_session, app):
        tz = app.config["HARVEST_TIMEZONE"]
        today = datetime.now(tz).date()
        start = datetime.combine(today, time.min, tzinfo=tz)
        utc_start = start.astimezone(timezone.utc)
        product = _ensure_product(db_session)

        w, a_ale = make_worker_with_assignment(db_session, name="Alejandro")
        _make_entry(
            db_session, w, a_ale, Decimal("10.000"),
            utc_start + timedelta(hours=8), product,
        )

        a_ale.ended_at = datetime.now(timezone.utc)
        db.session.flush()

        a_luis = WorkerAssignment(worker_id=w.id, person_name="Luis")
        db_session.add(a_luis)
        db.session.flush()

        _make_entry(
            db_session, w, a_luis, Decimal("5.000"),
            utc_start + timedelta(hours=12), product,
        )

        total_luis = get_daily_total(a_luis.id, operational_date=today, tz=tz)
        total_ale = get_daily_total(a_ale.id, operational_date=today, tz=tz)

        assert total_luis == Decimal("5.000")
        assert total_ale == Decimal("10.000")

    def test_two_people_same_slot_different_totals(self, db_session, app):
        tz = app.config["HARVEST_TIMEZONE"]
        today = datetime.now(tz).date()
        start = datetime.combine(today, time.min, tzinfo=tz)
        utc_start = start.astimezone(timezone.utc)
        product = _ensure_product(db_session)

        w, a1 = make_worker_with_assignment(db_session, name="Persona1")
        _make_entry(db_session, w, a1, Decimal("8.000"), utc_start + timedelta(hours=9), product)

        a1.ended_at = datetime.now(timezone.utc)
        db.session.flush()

        a2 = WorkerAssignment(worker_id=w.id, person_name="Persona2")
        db_session.add(a2)
        db_session.flush()

        _make_entry(db_session, w, a2, Decimal("4.000"), utc_start + timedelta(hours=14), product)

        total1 = get_daily_total(a1.id, operational_date=today, tz=tz)
        total2 = get_daily_total(a2.id, operational_date=today, tz=tz)

        assert total1 == Decimal("8.000")
        assert total2 == Decimal("4.000")


class TestHistoryWorkerIdVsAssignmentId:
    def test_summary_returns_both_ids(self, db_session, app):
        tz = app.config["HARVEST_TIMEZONE"]
        today = datetime.now(tz).date()
        start = datetime.combine(today, time.min, tzinfo=tz)
        utc_start = start.astimezone(timezone.utc)
        product = _ensure_product(db_session)

        w, a = make_worker_with_assignment(db_session, name="TestWorker")
        _make_entry(db_session, w, a, Decimal("3.000"), utc_start + timedelta(hours=10), product)

        result = get_daily_summary(today, tz=tz)
        assert len(result["workers"]) == 1
        worker_data = result["workers"][0]
        assert worker_data["worker_id"] == w.id
        assert worker_data["worker_assignment_id"] == a.id
        assert worker_data["slot_number"] == w.slot_number
        assert worker_data["name"] == "TestWorker"
        assert worker_data["barcode"] == w.barcode

    def test_detail_by_assignment_id(self, db_session, app):
        tz = app.config["HARVEST_TIMEZONE"]
        today = datetime.now(tz).date()
        start = datetime.combine(today, time.min, tzinfo=tz)
        utc_start = start.astimezone(timezone.utc)
        product = _ensure_product(db_session)

        w, a = make_worker_with_assignment(db_session, name="DetailWorker")
        _make_entry(db_session, w, a, Decimal("2.500"), utc_start + timedelta(hours=10), product)

        entries = get_worker_entries(a.id, today, tz=tz)
        assert len(entries) == 1
        assert entries[0].weight_kg == Decimal("2.500")


class TestExcelSeparatesAssignments:
    def test_two_assignments_same_slot_two_rows(self, db_session, app):
        tz = app.config["HARVEST_TIMEZONE"]
        today = datetime.now(tz).date()
        start = datetime.combine(today, time.min, tzinfo=tz)
        utc_start = start.astimezone(timezone.utc)
        product = _ensure_product(db_session)

        w, a1 = make_worker_with_assignment(db_session, name="Excel1")
        _make_entry(db_session, w, a1, Decimal("5.000"), utc_start + timedelta(hours=8), product)

        a1.ended_at = datetime.now(timezone.utc)
        db.session.flush()

        a2 = WorkerAssignment(worker_id=w.id, person_name="Excel2")
        db_session.add(a2)
        db_session.flush()

        _make_entry(db_session, w, a2, Decimal("3.000"), utc_start + timedelta(hours=12), product)

        xlsx_bytes, _ = generate_harvest_export(today, today, tz=tz)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Resumen"]

        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "TOTALES":
                break
            data_rows.append(row)

        assert len(data_rows) == 2
        names = [r[0] for r in data_rows]
        assert "Excel1" in names
        assert "Excel2" in names

    def test_movements_sheet_preserves_snapshots(self, db_session, app):
        tz = app.config["HARVEST_TIMEZONE"]
        today = datetime.now(tz).date()
        start = datetime.combine(today, time.min, tzinfo=tz)
        utc_start = start.astimezone(timezone.utc)
        product = _ensure_product(db_session)

        w, a = make_worker_with_assignment(db_session, name="SnapTest")
        _make_entry(db_session, w, a, Decimal("7.000"), utc_start + timedelta(hours=9), product)

        xlsx_bytes, _ = generate_harvest_export(today, today, tz=tz)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Movimientos"]

        row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        assert row[3] == "SnapTest"
        assert row[4] == w.barcode


class TestRegisterNoMovementWithoutAssignment:
    def test_unassigned_worker_returns_error(self, db_session, app):
        w = make_worker(db_session, name="Unassigned")

        with pytest.raises(WorkerUnassignedError):
            register_harvest(w.barcode, Decimal("5.000"), product_id=1)

        entries = HarvestEntry.query.filter_by(worker_id=w.id).all()
        assert len(entries) == 0


class TestAdminNoNPlusOne:
    def test_no_n_plus_one_on_list(self, admin_client, db_session, app):
        for slot in range(1, 151):
            existing = Worker.query.filter_by(slot_number=slot).first()
            if existing is None:
                w = Worker(barcode=f"TRB{slot:06d}", slot_number=slot, active=True)
                db_session.add(w)
        db_session.flush()

        assignment_selects = []

        def _track_sql(conn, cursor, stmt, parameters, context, executemany):
            stmt_upper = stmt.strip().upper()
            if stmt_upper.startswith("SELECT") and "WORKER_ASSIGNMENTS" in stmt_upper:
                assignment_selects.append(stmt)

        engine = db.engine
        from sqlalchemy import event
        event.listen(engine, "before_cursor_execute", _track_sql)

        try:
            resp = admin_client.get("/api/admin/worker-slots")
            assert resp.status_code == 200
            data = resp.get_json()
            assert len(data) == 150
        finally:
            event.remove(engine, "before_cursor_execute", _track_sql)

        assert len(assignment_selects) == 1, (
            f"Expected exactly 1 SELECT on worker_assignments (batch query), "
            f"got {len(assignment_selects)}: {assignment_selects}"
        )
