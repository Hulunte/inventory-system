import io
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from zoneinfo import ZoneInfo

from app.models.harvest_entry import HarvestEntry
from app.models.worker import Worker
from app.services.export_service import _safe_text, _to_local_naive
from app.services.report_service import date_range_to_utc, get_week_ranges


HARVEST_TZ = ZoneInfo("America/Mexico_City")

FIXED_NOW = datetime(2026, 6, 15, 14, 30, 0, tzinfo=timezone.utc)


class TestSafeText:
    @pytest.mark.parametrize("value,expected", [
        ("hello", "hello"),
        ("=1+1", "'=1+1"),
        ("+foo", "'+foo"),
        ("-bar", "'-bar"),
        ("@SUM(A1)", "'@SUM(A1)"),
        ("\tfoo", "'\tfoo"),
        ("\nfoo", "'\nfoo"),
        ("\rfoo", "'\rfoo"),
        ("'ABC", "'ABC"),
        ("", ""),
    ])
    def test_text_values(self, value, expected):
        assert _safe_text(value) == expected

    def test_non_string_returned_as_is(self):
        assert _safe_text(42) == 42
        assert _safe_text(None) is None

    def test_formula_cell_not_type_formula_after_roundtrip(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=_safe_text("=SUM(A1:A10)"))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb2 = load_workbook(buf)
        ws2 = wb2.active
        cell = ws2["A1"]
        assert cell.data_type != "f"


class TestToLocalNaive:
    def test_none_returns_none(self):
        assert _to_local_naive(None, HARVEST_TZ) is None

    def test_utc_to_local(self):
        utc_dt = datetime(2026, 1, 15, 10, 0, 0, tzinfo=timezone.utc)
        local = _to_local_naive(utc_dt, HARVEST_TZ)
        assert local.tzinfo is None
        assert local.hour == 4


class TestDateRangeToUtc:
    def test_utc_boundary_start_inclusive(self, admin_client, db_session):
        start_local = datetime(2026, 3, 10, 0, 0, 0, tzinfo=HARVEST_TZ)
        start_utc = start_local.astimezone(timezone.utc)

        with db_session.begin_nested():
            w = Worker(name="BoundaryStart", barcode="BST001")
            db_session.add(w)
            db_session.flush()
            entry = HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("1.000"),
                created_at=start_utc,
            )
            db_session.add(entry)
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-03-10&end_date=2026-03-10"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert entry.id in ids

    def test_utc_boundary_before_start_excluded(self, admin_client, db_session):
        start_local = datetime(2026, 3, 10, 0, 0, 0, tzinfo=HARVEST_TZ)
        start_utc = start_local.astimezone(timezone.utc)
        one_second_before = start_utc - timedelta(seconds=1)

        with db_session.begin_nested():
            w = Worker(name="BeforeStart", barcode="BST002")
            db_session.add(w)
            db_session.flush()
            entry = HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("1.000"),
                created_at=one_second_before,
            )
            db_session.add(entry)
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-03-10&end_date=2026-03-10"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert entry.id not in ids

    def test_utc_boundary_before_end_included(self, admin_client, db_session):
        end_local = datetime(2026, 3, 12, 0, 0, 0, tzinfo=HARVEST_TZ)
        end_utc = end_local.astimezone(timezone.utc)
        one_second_before_end = end_utc - timedelta(seconds=1)

        with db_session.begin_nested():
            w = Worker(name="BeforeEnd", barcode="BEN001")
            db_session.add(w)
            db_session.flush()
            entry = HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("1.000"),
                created_at=one_second_before_end,
            )
            db_session.add(entry)
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-03-10&end_date=2026-03-11"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert entry.id in ids

    def test_utc_boundary_at_end_excluded(self, admin_client, db_session):
        end_local = datetime(2026, 3, 12, 0, 0, 0, tzinfo=HARVEST_TZ)
        end_utc = end_local.astimezone(timezone.utc)

        with db_session.begin_nested():
            w = Worker(name="AtEnd", barcode="AEN001")
            db_session.add(w)
            db_session.flush()
            entry = HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("1.000"),
                created_at=end_utc,
            )
            db_session.add(entry)
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-03-10&end_date=2026-03-11"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert entry.id not in ids


class TestExportAuth:
    def test_unauthorized_returns_401(self, client):
        resp = client.get("/api/reports/harvest/export?start_date=2026-01-01&end_date=2026-01-31")
        assert resp.status_code == 401

    def test_admin_session_allows_download(self, admin_client):
        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-01-01&end_date=2026-01-05"
        )
        assert resp.status_code == 200
        assert "spreadsheetml.sheet" in resp.content_type


class TestExportSheets:
    def test_sheet_order(self, admin_client):
        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-01-01&end_date=2026-01-05"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        assert wb.sheetnames == ["Movimientos", "Resumen"]

    def test_filename_format(self, admin_client):
        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-03-10&end_date=2026-03-15"
        )
        cd = resp.headers["Content-Disposition"]
        assert "inventario_2026-03-10_a_2026-03-15.xlsx" in cd


class TestEmptyPeriod:
    def test_headers_and_structure(self, admin_client):
        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2020-01-01&end_date=2020-01-05"
        )
        wb = load_workbook(io.BytesIO(resp.data))

        ws_mov = wb["Movimientos"]
        assert ws_mov.max_row == 1
        assert ws_mov.freeze_panes == "A2"
        assert ws_mov.auto_filter.ref == "A1:I1"

        ws_res = wb["Resumen"]
        assert ws_res.max_row == 2
        assert ws_res.freeze_panes == "A2"
        assert ws_res.auto_filter.ref == "A1:F1"


class TestVigenteEntry:
    def test_active_entry_appears(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="Worker A", barcode="WA001")
            db_session.add(w)
            db_session.flush()
            entry = HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("45.500"),
                created_at=FIXED_NOW,
            )
            db_session.add(entry)
            db_session.flush()
            entry_id = entry.id

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == entry_id:
                row = r
                break
        assert row is not None
        assert ws.cell(row=row, column=7).value == "Vigente"
        assert ws.cell(row=row, column=9).value is None

        ws_res = wb["Resumen"]
        assert ws_res.cell(row=2, column=1).value == "Worker A"
        assert ws_res.cell(row=2, column=3).value == 1
        assert ws_res.cell(row=2, column=4).value == Decimal("45.500")


class TestAnuladoEntry:
    def test_voided_entry_shows_reason(self, admin_client, db_session):
        voided_at = datetime(2026, 6, 15, 10, 0, 0, tzinfo=timezone.utc)
        with db_session.begin_nested():
            w = Worker(name="Worker B", barcode="WB001")
            db_session.add(w)
            db_session.flush()
            entry = HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("12.000"),
                created_at=FIXED_NOW,
                voided=True, voided_at=voided_at,
                void_reason="Duplicate entry",
            )
            db_session.add(entry)
            db_session.flush()
            entry_id = entry.id

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == entry_id:
                row = r
                break
        assert row is not None
        assert ws.cell(row=row, column=7).value == "Anulado"
        assert ws.cell(row=row, column=9).value == "Duplicate entry"

        ws_res = wb["Resumen"]
        assert ws_res.cell(row=2, column=5).value == 1
        assert ws_res.cell(row=2, column=6).value == Decimal("12.000")


class TestInactiveWorker:
    def test_inactive_worker_included_in_export(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="Inactive W", barcode="INW001", active=False)
            db_session.add(w)
            db_session.flush()
            entry = HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("8.000"),
                created_at=FIXED_NOW,
            )
            db_session.add(entry)
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws_mov = wb["Movimientos"]
        names = [ws_mov.cell(row=r, column=4).value for r in range(2, ws_mov.max_row + 1)]
        assert "Inactive W" in names


class TestResumenTotals:
    def test_separate_vigente_anulado_totals(self, admin_client, db_session):
        voided_at = datetime(2026, 6, 15, 8, 0, 0, tzinfo=timezone.utc)
        with db_session.begin_nested():
            w = Worker(name="TotalWorker", barcode="TWT001")
            db_session.add(w)
            db_session.flush()
            db_session.add(HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("20.000"),
                created_at=FIXED_NOW,
            ))
            db_session.add(HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("15.500"),
                created_at=FIXED_NOW,
                voided=True, voided_at=voided_at, void_reason="Error",
            ))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Resumen"]

        worker_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "TotalWorker":
                worker_row = r
                break
        assert worker_row is not None
        assert ws.cell(row=worker_row, column=3).value == 1
        assert ws.cell(row=worker_row, column=4).value == Decimal("20.000")
        assert ws.cell(row=worker_row, column=5).value == 1
        assert ws.cell(row=worker_row, column=6).value == Decimal("15.500")

        total_row = ws.max_row
        assert ws.cell(row=total_row, column=1).value == "TOTALES"
        assert ws.cell(row=total_row, column=3).value == 1
        assert ws.cell(row=total_row, column=4).value == Decimal("20.000")
        assert ws.cell(row=total_row, column=5).value == 1
        assert ws.cell(row=total_row, column=6).value == Decimal("15.500")


class TestDecimalWeightPrecision:
    def test_exact_decimal_sum(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="DecWorker", barcode="DEC001")
            db_session.add(w)
            db_session.flush()
            db_session.add(HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("1.111"),
                created_at=FIXED_NOW,
            ))
            db_session.add(HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("2.222"),
                created_at=FIXED_NOW,
            ))
            db_session.add(HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("3.333"),
                created_at=FIXED_NOW,
            ))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws_mov = wb["Movimientos"]
        total = Decimal("0.000")
        for r in range(2, ws_mov.max_row + 1):
            cell = ws_mov.cell(row=r, column=6)
            assert cell.data_type == "n"
            assert cell.number_format == "0.000"
            total += Decimal(str(cell.value))
        assert total == Decimal("6.666")

        ws_res = wb["Resumen"]
        total_row = ws_res.max_row
        assert Decimal(str(ws_res.cell(row=total_row, column=4).value)).quantize(Decimal("0.001")) == Decimal("6.666")


class TestWeightCellFormat:
    def test_weight_cells_are_numeric(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="FmtWorker", barcode="FMT001")
            db_session.add(w)
            db_session.flush()
            db_session.add(HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("99.999"),
                created_at=FIXED_NOW,
            ))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        cell = ws.cell(row=2, column=6)
        assert cell.number_format == "0.000"
        assert isinstance(cell.value, (int, float, Decimal))


class TestDateAndTimeCellTypes:
    def test_date_and_time_are_excel_types_without_tzinfo(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="DTWorker", barcode="DT001")
            db_session.add(w)
            db_session.flush()
            entry = HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("5.000"),
                created_at=FIXED_NOW,
            )
            db_session.add(entry)
            db_session.flush()
            entry_id = entry.id

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == entry_id:
                row = r
                break
        assert row is not None

        date_cell = ws.cell(row=row, column=2)
        assert isinstance(date_cell.value, date)
        assert date_cell.number_format == "YYYY-MM-DD"

        time_cell = ws.cell(row=row, column=3)
        assert isinstance(time_cell.value, time)
        assert time_cell.number_format == "HH:MM:SS"

    def test_voided_at_is_datetime_without_tzinfo(self, admin_client, db_session):
        voided_at = datetime(2026, 6, 15, 12, 30, 0, tzinfo=timezone.utc)
        with db_session.begin_nested():
            w = Worker(name="VAWorker", barcode="VA001")
            db_session.add(w)
            db_session.flush()
            entry = HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("5.000"),
                created_at=FIXED_NOW,
                voided=True, voided_at=voided_at,
                void_reason="test",
            )
            db_session.add(entry)
            db_session.flush()
            entry_id = entry.id

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == entry_id:
                row = r
                break
        assert row is not None
        void_cell = ws.cell(row=row, column=8)
        assert isinstance(void_cell.value, datetime)
        assert void_cell.value.tzinfo is None


class TestFilterQ:
    def test_filter_by_name(self, admin_client, db_session):
        with db_session.begin_nested():
            w1 = Worker(name="XAlpha", barcode="XFA001")
            w2 = Worker(name="XBeta", barcode="XFB001")
            db_session.add_all([w1, w2])
            db_session.flush()
            db_session.add(HarvestEntry(worker_id=w1.id, weight_kg=Decimal("1.000"), created_at=FIXED_NOW))
            db_session.add(HarvestEntry(worker_id=w2.id, weight_kg=Decimal("2.000"), created_at=FIXED_NOW))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15&q=XAlpha"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        names = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
        assert "XAlpha" in names
        assert "XBeta" not in names

    def test_filter_by_barcode(self, admin_client, db_session):
        with db_session.begin_nested():
            w1 = Worker(name="XOne", barcode="XFX001")
            w2 = Worker(name="XTwo", barcode="XFY001")
            db_session.add_all([w1, w2])
            db_session.flush()
            db_session.add(HarvestEntry(worker_id=w1.id, weight_kg=Decimal("1.000"), created_at=FIXED_NOW))
            db_session.add(HarvestEntry(worker_id=w2.id, weight_kg=Decimal("2.000"), created_at=FIXED_NOW))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15&q=XFX001"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        names = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
        assert "XOne" in names
        assert "XTwo" not in names


class TestFormulaProtectionInColumns:
    def test_formula_in_name_column_is_escaped(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="=CMD('calc.exe')", barcode="FP001")
            db_session.add(w)
            db_session.flush()
            db_session.add(HarvestEntry(worker_id=w.id, weight_kg=Decimal("1.000"), created_at=FIXED_NOW))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        cell = ws.cell(row=2, column=4)
        assert cell.data_type != "f"

        ws_res = wb["Resumen"]
        res_cell = ws_res.cell(row=2, column=1)
        assert res_cell.data_type != "f"

    def test_formula_in_barcode_column_is_escaped(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="Safe", barcode="+remote_cmd")
            db_session.add(w)
            db_session.flush()
            db_session.add(HarvestEntry(worker_id=w.id, weight_kg=Decimal("1.000"), created_at=FIXED_NOW))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        cell = ws.cell(row=2, column=5)
        assert cell.data_type != "f"

        ws_res = wb["Resumen"]
        res_cell = ws_res.cell(row=2, column=2)
        assert res_cell.data_type != "f"

    def test_formula_in_void_reason_column_is_escaped(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="Safe", barcode="FP003")
            db_session.add(w)
            db_session.flush()
            db_session.add(HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("1.000"),
                created_at=FIXED_NOW,
                voided=True, voided_at=FIXED_NOW,
                void_reason="@SUM(A1:A10)",
            ))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        cell = ws.cell(row=2, column=9)
        assert cell.data_type != "f"


class TestLegitimateApostrophe:
    def test_apostrophe_in_name_preserved(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="O'Brien", barcode="AP001")
            db_session.add(w)
            db_session.flush()
            db_session.add(HarvestEntry(worker_id=w.id, weight_kg=Decimal("1.000"), created_at=FIXED_NOW))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        assert ws.cell(row=2, column=4).value == "O'Brien"

        ws_res = wb["Resumen"]
        assert ws_res.cell(row=2, column=1).value == "O'Brien"


class TestDateValidation:
    @pytest.mark.parametrize("params,expected_status", [
        ("start_date=not-a-date&end_date=2026-01-31", 400),
        ("start_date=2026-12-31&end_date=2026-01-01", 400),
        ("", 400),
    ])
    def test_invalid_dates(self, admin_client, params, expected_status):
        resp = admin_client.get(f"/api/reports/harvest/export?{params}")
        assert resp.status_code == expected_status


class TestEntryOrdering:
    def test_ordered_by_created_at_then_id(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="OrderW", barcode="ORD001")
            db_session.add(w)
            db_session.flush()
            ts_later = datetime(2026, 6, 15, 10, 0, 0, tzinfo=timezone.utc)
            ts_same = datetime(2026, 6, 15, 9, 0, 0, tzinfo=timezone.utc)
            e_later = HarvestEntry(worker_id=w.id, weight_kg=Decimal("3.000"), created_at=ts_later)
            e_first = HarvestEntry(worker_id=w.id, weight_kg=Decimal("1.000"), created_at=ts_same)
            e_second = HarvestEntry(worker_id=w.id, weight_kg=Decimal("2.000"), created_at=ts_same)
            db_session.add(e_later)
            db_session.flush()
            db_session.add_all([e_first, e_second])
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert ids[0] == e_first.id
        assert ids[1] == e_second.id
        assert ids[2] == e_later.id


class TestAutofilterAndFreeze:
    def test_freeze_always_a2(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="FreezeW", barcode="FRZ001")
            db_session.add(w)
            db_session.flush()
            db_session.add(HarvestEntry(worker_id=w.id, weight_kg=Decimal("1.000"), created_at=FIXED_NOW))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        assert wb["Movimientos"].freeze_panes == "A2"
        assert wb["Resumen"].freeze_panes == "A2"

    def test_resumen_autofilter_excludes_totals(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="AFWorker", barcode="AF001")
            db_session.add(w)
            db_session.flush()
            db_session.add(HarvestEntry(worker_id=w.id, weight_kg=Decimal("1.000"), created_at=FIXED_NOW))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-06-15&end_date=2026-06-15"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws_res = wb["Resumen"]
        assert ws_res.auto_filter.ref == "A1:F2"
        total_row = ws_res.max_row
        assert ws_res.cell(row=total_row, column=1).value == "TOTALES"
        assert str(total_row) not in ws_res.auto_filter.ref


class TestMovimientosHeaders:
    def test_header_names(self, admin_client):
        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-01-01&end_date=2026-01-05"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Movimientos"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert headers == [
            "ID", "Fecha", "Hora", "Trabajador", "Código",
            "Peso (kg)", "Estado", "Fecha y hora de anulación", "Motivo de anulación",
        ]


class TestResumenHeaders:
    def test_header_names(self, admin_client):
        resp = admin_client.get(
            "/api/reports/harvest/export?start_date=2026-01-01&end_date=2026-01-05"
        )
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Resumen"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == [
            "Trabajador", "Código", "Movimientos vigentes",
            "Peso vigente (kg)", "Movimientos anulados", "Peso anulado (kg)",
        ]


class TestGetWeekRanges:
    def test_current_week_monday_to_sunday(self):
        ref = date(2026, 6, 17)
        (mc, sc), (mp, sp) = get_week_ranges(ref)
        assert mc == date(2026, 6, 15)
        assert sc == date(2026, 6, 21)
        assert mp == date(2026, 6, 8)
        assert sp == date(2026, 6, 14)

    def test_monday_is_start_of_week(self):
        ref = date(2026, 6, 15)
        (mc, sc), (mp, sp) = get_week_ranges(ref)
        assert mc.weekday() == 0
        assert sc.weekday() == 6
        assert mp.weekday() == 0
        assert sp.weekday() == 6

    def test_weeks_are_consecutive(self):
        ref = date(2026, 6, 17)
        (mc, sc), (mp, sp) = get_week_ranges(ref)
        assert sp + timedelta(days=1) == mc


class TestReportButtonVisibility:
    def test_export_button_not_rendered_without_session(self, client):
        resp = client.get("/reports")
        assert resp.status_code == 200
        assert b"Exportar inventario" not in resp.data

    def test_export_button_rendered_with_admin_session(self, admin_client):
        resp = admin_client.get("/reports")
        assert resp.status_code == 200
        assert b"Exportar inventario" in resp.data


class TestExistingReportBehavior:
    def test_json_report_unchanged(self, admin_client, db_session):
        with db_session.begin_nested():
            w = Worker(name="JsonTest", barcode="JS001")
            db_session.add(w)
            db_session.flush()
            db_session.add(HarvestEntry(
                worker_id=w.id, weight_kg=Decimal("30.000"),
                created_at=FIXED_NOW,
            ))
            db_session.flush()

        resp = admin_client.get(
            "/api/reports/harvest?start_date=2026-06-15&end_date=2026-06-15"
        )
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["start_date"] == "2026-06-15"
        assert data["end_date"] == "2026-06-15"
        assert len(data["workers"]) == 1
        assert data["workers"][0]["name"] == "JsonTest"
        assert data["summary"]["total_entries"] == 1
        assert data["summary"]["total_weight_kg"] == "30.000"
